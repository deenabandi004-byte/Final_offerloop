"""
Contact import routes - Upload CSV/Excel to import contacts
"""
import csv
import io
import requests
from datetime import datetime
from flask import Blueprint, request, jsonify, Response

from ..extensions import require_firebase_auth
from ..extensions import get_db
from app.utils.exceptions import OfferloopException, ValidationError
from app.config import PDL_BASE_URL, PEOPLE_DATA_LABS_API_KEY
from app.routes.linkedin_import import (
    resolve_email_for_linkedin_import,
    extract_contact_from_pdl_person_enhanced,
    normalize_linkedin_url,
)
from app.services.reply_generation import batch_generate_emails
from app.services.gmail_client import create_gmail_draft_for_user, download_resume_from_url

contact_import_bp = Blueprint('contact_import', __name__, url_prefix='/api/contacts')

# Cost per imported contact
CREDITS_PER_CONTACT = 15

# Max contacts to enrich via PDL per import (LinkedIn URL -> email lookup)
ENRICHMENT_CAP = 50

# Column mapping - maps common CSV header variations to our schema
COLUMN_MAPPINGS = {
    'firstName': ['firstname', 'first_name', 'first name', 'fname', 'given name'],
    'lastName': ['lastname', 'last_name', 'last name', 'lname', 'surname', 'family name'],
    'email': ['email', 'email address', 'e-mail', 'mail', 'work email', 'personal email'],
    'linkedinUrl': ['linkedin', 'linkedinurl', 'linkedin_url', 'linkedin url', 'linkedin profile', 'profile url'],
    'company': ['company', 'company name', 'organization', 'employer', 'firm', 'workplace'],
    'jobTitle': ['jobtitle', 'job_title', 'job title', 'title', 'role', 'position'],
    'college': ['college', 'university', 'school', 'education', 'alma mater'],
    'location': ['location', 'city', 'address', 'city, state', 'city/state'],
    'city': ['city'],
    'state': ['state', 'province', 'region'],
    'phone': ['phone', 'phone number', 'mobile', 'cell', 'telephone'],
}


def normalize_header(header: str) -> str:
    """Normalize a header string for matching"""
    return header.lower().strip().replace('-', ' ').replace('_', ' ')


def map_columns(headers: list) -> dict:
    """
    Map CSV headers to our schema fields.
    Returns a dict of {csv_index: our_field_name}
    """
    mapping = {}
    normalized_headers = [normalize_header(h) for h in headers]
    
    for our_field, variations in COLUMN_MAPPINGS.items():
        for idx, header in enumerate(normalized_headers):
            if header in variations or header == our_field.lower():
                mapping[idx] = our_field
                break
    
    return mapping


def parse_csv_content(file_content: str) -> tuple:
    """Parse CSV content and return headers and rows"""
    reader = csv.reader(io.StringIO(file_content))
    rows = list(reader)
    
    if not rows:
        raise ValidationError("CSV file is empty", field="file")
    
    headers = rows[0]
    data_rows = rows[1:]
    
    return headers, data_rows


def parse_row_to_contact(row: list, column_mapping: dict, headers: list) -> dict:
    """Convert a CSV row to a contact dict using the column mapping"""
    contact = {}
    
    for idx, value in enumerate(row):
        if idx in column_mapping:
            field = column_mapping[idx]
            contact[field] = value.strip() if value else ''
    
    # Handle city + state -> location
    if 'city' in contact or 'state' in contact:
        city = contact.pop('city', '')
        state = contact.pop('state', '')
        if not contact.get('location'):
            contact['location'] = ', '.join(filter(None, [city, state]))
    
    return contact


def _is_valid_contact(contact: dict) -> bool:
    """True if contact has at least name, email, or LinkedIn URL."""
    has_name = contact.get('firstName') and contact.get('lastName')
    has_email = bool(contact.get('email', '').strip())
    has_linkedin = bool(contact.get('linkedinUrl', '').strip())
    return bool(has_name or has_email or has_linkedin)


@contact_import_bp.route('/import/preview', methods=['POST'])
@require_firebase_auth
def preview_import():
    """
    Preview CSV import - parse file and return column mapping suggestions.
    Does not deduct credits or save contacts.
    """
    try:
        db = get_db()
        user_id = request.firebase_user['uid']
        
        if not db:
            raise OfferloopException("Database not initialized", error_code="DATABASE_ERROR")
        
        # Check user tier - must be pro or above
        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            raise OfferloopException("User not found", error_code="USER_NOT_FOUND")
        
        user_data = user_doc.to_dict()
        tier = user_data.get('tier', 'free')
        
        # TODO: Re-enable tier restriction after testing
        # if tier == 'free':
        #     return jsonify({
        #         'error': 'Contact import is available for Pro and Elite users only',
        #         'upgrade_required': True
        #     }), 403
        
        # Get file from request
        if 'file' not in request.files:
            raise ValidationError("No file provided", field="file")
        
        file = request.files['file']
        
        if not file.filename:
            raise ValidationError("No file selected", field="file")
        
        # Check file extension
        filename = file.filename.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            raise ValidationError("File must be CSV or Excel (.csv, .xlsx, .xls)", field="file")
        
        # Parse file
        if filename.endswith('.csv'):
            file_content = file.read().decode('utf-8-sig')  # Handle BOM
            headers, data_rows = parse_csv_content(file_content)
        else:
            # Handle Excel files
            try:
                import openpyxl
                from io import BytesIO
                
                workbook = openpyxl.load_workbook(BytesIO(file.read()), read_only=True)
                sheet = workbook.active
                
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    raise ValidationError("Excel file is empty", field="file")
                
                headers = [str(h) if h else '' for h in rows[0]]
                data_rows = [[str(cell) if cell else '' for cell in row] for row in rows[1:]]
                
            except ImportError:
                raise OfferloopException(
                    "Excel support not available. Please upload a CSV file.",
                    error_code="EXCEL_NOT_SUPPORTED"
                )
        
        # Auto-detect column mappings
        column_mapping = map_columns(headers)
        
        # Get user's current credits
        credits = user_data.get('credits', 0)
        
        # Count valid rows and enrichment stats
        valid_rows = []
        contacts_with_email = 0
        contacts_needing_enrichment = 0
        contacts_unenrichable = 0
        for row in data_rows:
            contact = parse_row_to_contact(row, column_mapping, headers)
            if not _is_valid_contact(contact):
                continue
            valid_rows.append(contact)
            email = (contact.get('email') or '').strip()
            linkedin = (contact.get('linkedinUrl') or '').strip()
            if email:
                contacts_with_email += 1
            elif linkedin:
                contacts_needing_enrichment += 1
            else:
                contacts_unenrichable += 1
        
        # Calculate credit cost
        total_cost = len(valid_rows) * CREDITS_PER_CONTACT
        can_afford = credits >= total_cost
        max_affordable = credits // CREDITS_PER_CONTACT
        
        # Return preview with first 5 parsed contacts as sample
        return jsonify({
            'success': True,
            'headers': headers,
            'column_mapping': {str(k): v for k, v in column_mapping.items()},
            'unmapped_headers': [h for idx, h in enumerate(headers) if idx not in column_mapping],
            'total_rows': len(data_rows),
            'valid_rows': len(valid_rows),
            'sample_contacts': valid_rows[:5],
            'credits': {
                'available': credits,
                'cost_per_contact': CREDITS_PER_CONTACT,
                'total_cost': total_cost,
                'can_afford': can_afford,
                'max_affordable': max_affordable
            },
            'enrichment': {
                'contacts_with_email': contacts_with_email,
                'contacts_needing_enrichment': min(contacts_needing_enrichment, ENRICHMENT_CAP),
                'contacts_needing_enrichment_total': contacts_needing_enrichment,
                'contacts_unenrichable': contacts_unenrichable,
                'enrichment_cap': ENRICHMENT_CAP,
            }
        })
        
    except ValidationError as ve:
        return jsonify({'error': ve.message, 'field': getattr(ve, 'field', None)}), 400
    except OfferloopException as oe:
        return jsonify({'error': oe.message, 'error_code': oe.error_code}), 500
    except Exception as e:
        print(f"Error previewing import: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to preview import: {str(e)}'}), 500


@contact_import_bp.route('/import', methods=['POST'])
@require_firebase_auth
def import_contacts():
    """
    Import contacts from CSV/Excel file.
    Deducts 15 credits per contact imported.
    Skips duplicates (by email or LinkedIn URL).
    """
    try:
        db = get_db()
        user_id = request.firebase_user['uid']
        
        if not db:
            raise OfferloopException("Database not initialized", error_code="DATABASE_ERROR")
        
        # Check user tier - must be pro or above
        user_ref = db.collection('users').document(user_id)
        user_doc = user_ref.get()
        
        if not user_doc.exists:
            raise OfferloopException("User not found", error_code="USER_NOT_FOUND")
        
        user_data = user_doc.to_dict()
        tier = user_data.get('tier', 'free')
        
        # TODO: Re-enable tier restriction after testing
        # if tier == 'free':
        #     return jsonify({
        #         'error': 'Contact import is available for Pro and Elite users only',
        #         'upgrade_required': True
        #     }), 403
        
        # Get current credits
        credits = user_data.get('credits', 0)
        
        # Get file from request
        if 'file' not in request.files:
            raise ValidationError("No file provided", field="file")
        
        file = request.files['file']
        
        if not file.filename:
            raise ValidationError("No file selected", field="file")
        
        # Check file extension
        filename = file.filename.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
            raise ValidationError("File must be CSV or Excel (.csv, .xlsx, .xls)", field="file")
        
        # Get optional custom column mapping from request
        custom_mapping = request.form.get('column_mapping')
        if custom_mapping:
            import json
            custom_mapping = json.loads(custom_mapping)
            # Convert string keys back to int
            custom_mapping = {int(k): v for k, v in custom_mapping.items()}
        
        # Parse file
        if filename.endswith('.csv'):
            file_content = file.read().decode('utf-8-sig')
            headers, data_rows = parse_csv_content(file_content)
        else:
            try:
                import openpyxl
                from io import BytesIO
                
                workbook = openpyxl.load_workbook(BytesIO(file.read()), read_only=True)
                sheet = workbook.active
                
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    raise ValidationError("Excel file is empty", field="file")
                
                headers = [str(h) if h else '' for h in rows[0]]
                data_rows = [[str(cell) if cell else '' for cell in row] for row in rows[1:]]
                
            except ImportError:
                raise OfferloopException(
                    "Excel support not available. Please upload a CSV file.",
                    error_code="EXCEL_NOT_SUPPORTED"
                )
        
        # Use custom mapping or auto-detect
        column_mapping = custom_mapping if custom_mapping else map_columns(headers)
        
        contacts_ref = db.collection('users').document(user_id).collection('contacts')
        today = datetime.now().strftime('%m/%d/%Y')
        
        created = 0
        skipped_duplicate = 0
        skipped_invalid = 0
        skipped_no_credits = 0
        created_contacts = []
        enriched_count = 0
        enrichment_failed = 0
        enrichment_capped = 0
        contacts_for_drafting = []
        
        for row in data_rows:
            # Parse row to contact
            contact_data = parse_row_to_contact(row, column_mapping, headers)
            
            first_name = contact_data.get('firstName', '').strip()
            last_name = contact_data.get('lastName', '').strip()
            email = contact_data.get('email', '').strip()
            linkedin = contact_data.get('linkedinUrl', '').strip()
            
            # Validate - must have name OR email OR linkedin
            if not _is_valid_contact(contact_data):
                skipped_invalid += 1
                continue
            
            # Check credits before processing
            if credits < CREDITS_PER_CONTACT:
                skipped_no_credits += 1
                continue
            
            # Enrichment: PDL lookup for rows with LinkedIn URL but no email (cap 50)
            if not email and linkedin:
                if enriched_count < ENRICHMENT_CAP:
                    pdl_url = normalize_linkedin_url(linkedin) or linkedin
                    if pdl_url and not pdl_url.startswith('http'):
                        pdl_url = f'https://www.linkedin.com/in/{pdl_url.split("/in/")[-1].rstrip("/")}' if '/in/' in pdl_url else None
                    if pdl_url and 'linkedin.com' in pdl_url:
                        try:
                            response = requests.get(
                                f"{PDL_BASE_URL}/person/enrich",
                                params={
                                    'api_key': PEOPLE_DATA_LABS_API_KEY,
                                    'profile': pdl_url,
                                    'pretty': True
                                },
                                timeout=30
                            )
                            if response.status_code == 200:
                                person_json = response.json()
                                person_data = person_json.get('data') if isinstance(person_json, dict) else None
                                if person_data:
                                    pdl_contact = extract_contact_from_pdl_person_enhanced(person_data)
                                    if pdl_contact:
                                        email_result = resolve_email_for_linkedin_import(pdl_contact, person_data)
                                        resolved_email = (email_result.get('email') or '').strip()
                                        if resolved_email and '@' in resolved_email:
                                            contact_data['email'] = resolved_email
                                            contact_data['emailSource'] = 'enriched'
                                            enriched_count += 1
                                        else:
                                            enrichment_failed += 1
                                    else:
                                        enrichment_failed += 1
                                else:
                                    enrichment_failed += 1
                            else:
                                enrichment_failed += 1
                                if response.status_code == 402:
                                    print(f"[ContactImport] PDL 402 quota exceeded for row")
                        except Exception as e:
                            enrichment_failed += 1
                            print(f"[ContactImport] Enrichment error for LinkedIn {linkedin[:50]!r}: {e}")
                    else:
                        enrichment_failed += 1
                else:
                    enrichment_capped += 1
            
            email = contact_data.get('email', '').strip()
            
            # Duplicate detection (after enrichment, so enriched email can match existing)
            is_duplicate = False
            if email:
                email_query = contacts_ref.where('email', '==', email).limit(1)
                if list(email_query.stream()):
                    is_duplicate = True
            if not is_duplicate and linkedin:
                linkedin_query = contacts_ref.where('linkedinUrl', '==', linkedin).limit(1)
                if list(linkedin_query.stream()):
                    is_duplicate = True
            if not is_duplicate and first_name and last_name and contact_data.get('company'):
                name_company_query = (contacts_ref
                    .where('firstName', '==', first_name)
                    .where('lastName', '==', last_name)
                    .where('company', '==', contact_data.get('company'))
                    .limit(1))
                if list(name_company_query.stream()):
                    is_duplicate = True
            
            if is_duplicate:
                skipped_duplicate += 1
                continue
            
            # Create contact (same schema as existing contacts)
            contact = {
                'firstName': first_name,
                'lastName': last_name,
                'email': email,
                'linkedinUrl': linkedin,
                'company': contact_data.get('company', ''),
                'jobTitle': contact_data.get('jobTitle', ''),
                'college': contact_data.get('college', ''),
                'location': contact_data.get('location', ''),
                'phone': contact_data.get('phone', ''),
                'firstContactDate': today,
                'status': 'Not Contacted',
                'lastContactDate': today,
                'userId': user_id,
                'createdAt': today,
                'importedAt': datetime.now().isoformat(),
                'importSource': 'spreadsheet',
                'emailSource': contact_data.get('emailSource', 'imported'),
            }
            
            doc_ref = contacts_ref.add(contact)
            doc_id = doc_ref[1].id
            contact['id'] = doc_id
            created_contacts.append(contact)
            created += 1
            credits -= CREDITS_PER_CONTACT
            
            if email:
                contacts_for_drafting.append({'doc_id': doc_id, 'contact': contact})
        
        # Update user's credits in database
        if created > 0:
            user_ref.update({
                'credits': credits,
                'lastCreditUsage': datetime.now().isoformat()
            })
        
        # Phase 2: Batch generate emails + create Gmail drafts
        drafts_created = 0
        try:
            if contacts_for_drafting:
                user_doc = user_ref.get()
                user_data_after = user_doc.to_dict() if user_doc.exists else {}
                resume_text = (user_data_after.get('resumeText') or '').strip()
                user_profile = {
                    'name': user_data_after.get('name', ''),
                    'email': request.firebase_user.get('email', ''),
                    'university': user_data_after.get('university', ''),
                    'major': user_data_after.get('major', ''),
                    'year': user_data_after.get('year', ''),
                }
                career_interests = user_data_after.get('careerInterests') or []
                if isinstance(career_interests, str):
                    career_interests = [career_interests] if career_interests else []
                
                email_contacts = []
                for item in contacts_for_drafting:
                    c = item['contact']
                    email_contacts.append({
                        'FirstName': c.get('firstName', ''),
                        'LastName': c.get('lastName', ''),
                        'Company': c.get('company', ''),
                        'Title': c.get('jobTitle', ''),
                        'Email': c.get('email', ''),
                        'LinkedIn': c.get('linkedinUrl', ''),
                    })
                
                email_results = {}
                try:
                    email_results = batch_generate_emails(
                        contacts=email_contacts,
                        resume_text=resume_text or None,
                        user_profile=user_profile,
                        career_interests=career_interests,
                        fit_context=None,
                        email_template_purpose='networking',
                    )
                except Exception as e:
                    print(f"[ContactImport] Email generation failed: {e}")
                    import traceback
                    traceback.print_exc()
                
                user_email = request.firebase_user.get('email')
                user_info = {
                    'name': user_profile.get('name', ''),
                    'email': user_profile.get('email', ''),
                    'phone': user_data_after.get('phone', ''),
                    'linkedin': user_data_after.get('linkedin', ''),
                }
                resume_content = None
                resume_filename = None
                resume_url = user_data_after.get('resumeUrl')
                if resume_url:
                    try:
                        resume_content, resume_filename = download_resume_from_url(resume_url)
                        resume_filename = resume_filename or user_data_after.get('resumeFileName') or 'resume.pdf'
                    except Exception as e:
                        print(f"[ContactImport] Resume download failed: {e}")
                
                for idx, item in enumerate(contacts_for_drafting):
                    r = email_results.get(idx) or email_results.get(str(idx))
                    if not r or not isinstance(r, dict):
                        continue
                    subject = (r.get('subject') or '').strip()
                    body = (r.get('body') or '').strip()
                    if not subject or not body:
                        continue
                    
                    contact_ref = db.collection('users').document(user_id).collection('contacts').document(item['doc_id'])
                    update_data = {
                        'emailSubject': subject,
                        'emailBody': body,
                        'draftCreatedAt': datetime.now().isoformat(),
                    }
                    
                    contact_for_draft = {
                        'FirstName': item['contact'].get('firstName', ''),
                        'LastName': item['contact'].get('lastName', ''),
                        'Email': item['contact'].get('email', ''),
                    }
                    try:
                        draft_result = create_gmail_draft_for_user(
                            contact=contact_for_draft,
                            email_subject=subject,
                            email_body=body,
                            tier='free',
                            user_email=user_email,
                            user_id=user_id,
                            user_info=user_info,
                            resume_content=resume_content,
                            resume_filename=resume_filename,
                        )
                        if draft_result and isinstance(draft_result, dict):
                            update_data['gmailDraftId'] = draft_result.get('draft_id', '')
                            update_data['gmailDraftUrl'] = draft_result.get('draft_url', '')
                            if draft_result.get('message_id'):
                                update_data['gmailMessageId'] = draft_result.get('message_id', '')
                            drafts_created += 1
                    except Exception as e:
                        print(f"[ContactImport] Gmail draft failed for contact {item['doc_id']}: {e}")
                    
                    contact_ref.update(update_data)
        except Exception as e:
            print(f"[ContactImport] Enrichment/draft phase error: {e}")
            import traceback
            traceback.print_exc()
        
        return jsonify({
            'success': True,
            'created': created,
            'skipped': {
                'duplicate': skipped_duplicate,
                'invalid': skipped_invalid,
                'no_credits': skipped_no_credits,
                'total': skipped_duplicate + skipped_invalid + skipped_no_credits
            },
            'enrichment': {
                'enriched': enriched_count,
                'failed': enrichment_failed,
                'capped': enrichment_capped,
            },
            'drafts': {
                'created': drafts_created,
                'total_eligible': len(contacts_for_drafting),
            },
            'contacts': created_contacts,
            'credits': {
                'spent': created * CREDITS_PER_CONTACT,
                'remaining': credits
            }
        })
        
    except ValidationError as ve:
        return jsonify({'error': ve.message, 'field': getattr(ve, 'field', None)}), 400
    except OfferloopException as oe:
        return jsonify({'error': oe.message, 'error_code': oe.error_code}), 500
    except Exception as e:
        print(f"Error importing contacts: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Failed to import contacts: {str(e)}'}), 500


@contact_import_bp.route('/import/template', methods=['GET'])
@require_firebase_auth
def download_template():
    """Download a CSV template for contact import"""
    headers = [
        'First Name',
        'Last Name', 
        'Email',
        'LinkedIn URL',
        'Company',
        'Job Title',
        'College',
        'Location',
        'Phone'
    ]
    
    sample_row = [
        'John',
        'Doe',
        'john.doe@example.com',
        'https://linkedin.com/in/johndoe',
        'Acme Corp',
        'Software Engineer',
        'MIT',
        'San Francisco, CA',
        '555-123-4567'
    ]
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(sample_row)
    
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=contact_import_template.csv'}
    )
